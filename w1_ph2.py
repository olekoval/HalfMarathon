import halfmarathon
import openpyxl

alldist = 60 # недельный обЪем, км
time_long = 120 # минут длительного бега в Длительную тренировку
long_pace = "6:20"  # длительн темп
repeats_1_plan = ((200, 200), (200, 400), (400, 200))
repeats_2_plan = ((200, 200),)
threshold_interval = 1.6 # дистанция порогового интервала, км
threshold_pace = "5:19" # пороговый темп


p = halfmarathon.Phase2(alldist)

# Параметр Длительной тренировки
distance_long = p.calculate_distance_long(time_long, long_pace)

# Параметры первой тренировки Пв (вместе с П-темпом)
rep1 = p.number_episodes_repeats(repeats_1_plan, .75) # Пв-темп
plan_rep1 = rep1[2]
distance_rep1 = rep1[1]

# Параметры второй тренировки Пв (вместе с П-темпом)
th = p.number_episodes_threshold(threshold_interval, threshold_pace) # П-темп
rep2 = p.number_episodes_repeats(repeats_2_plan, .25) # Пв-темп
plan_th_rep2 = f"{th[2]} + {rep2[2]}" # План тренировки
distance_th_rep2 = th[1] + rep2[1] # Общая дистанция тренировки

# Параметри легких тренировок
distance_easy = round((alldist - (distance_long + distance_rep1 + distance_th_rep2)) / 4, 2)


wb = openpyxl.Workbook()
ws = wb.active

name_day = ['Пн','Вт','Ср','Чт','Пт','Сб','Вс']
for i in range(7):
    ws['A'+ str(i + 2)] = name_day[i]

    
ws['A1'] = "День недели"
ws['B1'] = "Дистанция тренировки"
ws['C1'] = "План тренировки"

for i in [2, 4, 5, 7]:
    ws['B' + str(i)] = distance_easy
    if i < 7:
        ws['C' + str(i)] = "Л"
    else:
        ws['C' + str(i)] = "Л + 6 кб"

ws['B3'] = distance_rep1
ws['C3'] = plan_rep1

ws['B6'] = distance_th_rep2
ws['C6'] = plan_th_rep2

ws['B8'] = distance_long
ws['C8'] = "Л"

wb.save("w1.xlsx")
print("OK")
    


